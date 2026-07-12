from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT / "source" / "hakedis.xlsx"
FALLBACK_XLSX = ROOT / "source" / "teklif.xlsx"
MAP_IMAGE = ROOT / "public" / "vaziyet-plani.png"
OUTPUT_JSON = ROOT / "src" / "data" / "siteData.json"
MAP_WIDTH = 9362
MAP_HEIGHT = 6623


def slugify(value: str) -> str:
    table = str.maketrans(
        "çğıöşüÇĞİÖŞÜı",
        "cgiosuCGIOSUi",
    )
    value = str(value).translate(table).lower()
    value = re.sub(r"[^a-z0-9]+", "_", value).strip("_")
    return value or "is_kalemi"


def as_number(value):
    if value in (None, ""):
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


WORK_CATEGORY_BY_KEY = {
    "grup_sayisi": "sihhi",
    "dalgic_pompa": "sihhi",
    "yag_tutucu": "sihhi",
    "vrf_drenaj": "sihhi",
    "ara_istasyon": "sihhi",
    "karot_deligi": "sihhi",
    "petek": "isitma",
    "kollektor": "isitma",
    "sprink": "yangin",
    "yangin_dolabi": "yangin",
    "i_b_a": "yangin",
}

CATEGORY_WEIGHTS = {
    "sihhi": 34,
    "isitma": 33,
    "yangin": 33,
}


def work_category(key: str) -> str:
    return WORK_CATEGORY_BY_KEY.get(key, "sihhi")


GROUP_SOURCE_COLUMNS = {
    "sihhi_tesisat": (6, 8),
    "karot": (16,),
    "vrf": (9,),
    "pis_su_pompasi": (7,),
    "isitma_tesisati": (10, 11),
    "yangin_tesisati": (12, 13, 14),
    "ara_istasyon": (15,),
}


def make_key(value: str) -> str:
    value = str(value).translate(str.maketrans({"ı": "i", "İ": "i"}))
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^a-zA-Z0-9]+", "_", value).strip("_").lower()
    return value or "is_kalemi"


def find_sheet(workbook, needle, fallback_index):
    needle = make_key(needle)
    for sheet in workbook.worksheets:
        if needle in make_key(sheet.title):
            return sheet
    return workbook.worksheets[fallback_index]


def read_hakedis_works(workbook):
    sheet = find_sheet(workbook, "hakedis", 2)
    works_by_group = {}
    work_items_by_key = {}
    current_group = ""
    current_group_key = ""
    order = 0

    for row_index in range(1, sheet.max_row + 1):
        group = sheet.cell(row_index, 2).value
        name = sheet.cell(row_index, 3).value
        percent = as_number(sheet.cell(row_index, 7).value)
        if group:
            current_group = str(group).strip()
            current_group_key = make_key(current_group)
        if not name or percent <= 0 or not current_group_key:
            continue

        label = str(name).strip()
        if "toplam" in make_key(label):
            continue
        order += 1
        key = f"{current_group_key}_{make_key(label)}"
        work = {
            "key": key,
            "label": label,
            "quantity": 100,
            "weight": round(percent * 100, 2),
            "category": current_group_key,
            "hakedisPercent": round(percent * 100, 2),
            "unit": "percent",
            "order": order,
        }
        works_by_group.setdefault(current_group_key, []).append(work)
        work_items_by_key[key] = {
            "key": key,
            "label": label,
            "category": current_group_key,
            "weight": work["weight"],
            "unit": "percent",
            "order": order,
        }

    return works_by_group, list(sorted(work_items_by_key.values(), key=lambda item: item["order"]))


def building_active_group_keys(sheet, row_index):
    active = []
    for group_key, columns in GROUP_SOURCE_COLUMNS.items():
        if any(as_number(sheet.cell(row_index, column).value) > 0 for column in columns):
            active.append(group_key)
    return active


def read_buildings():
    source_path = SOURCE_XLSX if SOURCE_XLSX.exists() else FALLBACK_XLSX
    wb = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    ws = find_sheet(wb, "binalar", 0)
    works_by_group, work_items = read_hakedis_works(wb)
    buildings = []

    for row_index in range(3, ws.max_row + 1):
        code = ws.cell(row_index, 2).value
        name = ws.cell(row_index, 4).value
        if not code or not name:
            continue

        code = str(code).strip()
        works = []
        progress = {}
        for group_key in building_active_group_keys(ws, row_index):
            for work in works_by_group.get(group_key, []):
                works.append(dict(work))
                progress[work["key"]] = 0

        buildings.append(
            {
                "id": code,
                "code": code,
                "name": str(name).strip(),
                "lineColor": str(ws.cell(row_index, 3).value or "BELIRSIZ").strip(),
                "quantity": as_number(ws.cell(row_index, 5).value),
                "works": works,
                "progress": progress,
            }
        )

    return buildings, work_items


def make_users(buildings, work_items):
    all_ids = [building["id"] for building in buildings]
    all_work_keys = [work["key"] for work in work_items]
    red_ids = [building["id"] for building in buildings if building["lineColor"].upper() == "KIRMIZI"]
    turquoise_ids = [building["id"] for building in buildings if building["lineColor"].upper() == "TURKUAZ"]
    purple_ids = [
        building["id"]
        for building in buildings
        if building["lineColor"].upper() in {"MOR", "MAGENTA", "MAVİ"}
    ]

    return [
        {
            "id": "u-admin",
            "name": "Süper Admin",
            "username": "admin",
            "password": "admin123",
            "role": "admin",
            "permissions": all_ids,
            "workPermissions": all_work_keys,
        },
        {
            "id": "u-formen-kirmizi",
            "name": "Kırmızı Hat Formeni",
            "username": "formen1",
            "password": "formen123",
            "role": "foreman",
            "permissions": red_ids[:60],
            "workPermissions": [],
        },
        {
            "id": "u-formen-turkuaz",
            "name": "Turkuaz Hat Formeni",
            "username": "formen2",
            "password": "formen123",
            "role": "foreman",
            "permissions": turquoise_ids[:60],
            "workPermissions": [],
        },
        {
            "id": "u-formen-mor",
            "name": "Mor/Magenta Hat Formeni",
            "username": "formen3",
            "password": "formen123",
            "role": "foreman",
            "permissions": purple_ids[:60],
            "workPermissions": [],
        },
    ]


def main():
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Excel bulunamadi: {SOURCE_XLSX}")
    if not MAP_IMAGE.exists():
        raise FileNotFoundError(f"Map gorseli bulunamadi: {MAP_IMAGE}")

    buildings, work_items = read_buildings()
    data = {
        "map": {
            "image": "/vaziyet-plani.png",
            "width": MAP_WIDTH,
            "height": MAP_HEIGHT,
        },
        "buildings": buildings,
        "workItems": work_items,
        "users": make_users(buildings, work_items),
        "requests": [],
        "progressRanges": [
            {"id": "range-0-20", "min": 0, "max": 20, "color": "#ff4040", "label": "0-20"},
            {"id": "range-20-40", "min": 20, "max": 40, "color": "#ffac5a", "label": "20-40"},
            {"id": "range-40-60", "min": 40, "max": 60, "color": "#ffff3b", "label": "40-60"},
            {"id": "range-60-80", "min": 60, "max": 80, "color": "#91ff8f", "label": "60-80"},
            {"id": "range-80-100", "min": 80, "max": 100, "color": "#00f75c", "label": "80-100"},
        ],
    }

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        f"Generated {len(buildings)} buildings and {len(work_items)} work item types -> {OUTPUT_JSON}"
    )


if __name__ == "__main__":
    main()
