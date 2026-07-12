import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


CATEGORY_CONFIG = {
    "SIHHI TESISAT": {
        "id": "sihhi_tesisat",
        "label": "Sıhhi Tesisat",
        "driverKey": "sihhiGrupSayisi",
        "driverLabel": "Sıhhi grup",
        "automaticWeight": True,
    },
    "KAROT": {
        "id": "karot",
        "label": "Karot",
        "driverKey": "karotDeligiSayisi",
        "driverLabel": "Karot deliği",
        "automaticWeight": False,
    },
    "VRF": {
        "id": "vrf",
        "label": "VRF",
        "driverKey": "vrfDrenajMetraji",
        "driverLabel": "VRF drenaj",
        "automaticWeight": False,
    },
    "PIS SU POMPASI": {
        "id": "pis_su_pompasi",
        "label": "Pis Su Pompası",
        "driverKey": "dalgicPompaSayisi",
        "driverLabel": "Dalgıç pompa",
        "automaticWeight": False,
    },
    "ISITMA TESISATI": {
        "id": "isitma_tesisati",
        "label": "Isıtma Tesisatı",
        "driverKey": "petekSayisi",
        "driverLabel": "Petek",
        "automaticWeight": True,
    },
    "YANGIN TESISATI": {
        "id": "yangin_tesisati",
        "label": "Yangın Tesisatı",
        "driverKey": "sprinkSayisi",
        "driverLabel": "Sprink",
        "automaticWeight": True,
    },
    "ARA ISTASYON": {
        "id": "ara_istasyon",
        "label": "Ara İstasyon",
        "driverKey": "araIstasyonMetraji",
        "driverLabel": "Ara istasyon",
        "automaticWeight": False,
    },
}

QUANTITY_FIELDS = [
    ("sihhiGrupSayisi", "Sıhhi grup", 5),
    ("dalgicPompaSayisi", "Dalgıç pompa", 6),
    ("yagTutucuSayisi", "Yağ tutucu", 7),
    ("vrfDrenajMetraji", "VRF drenaj", 8),
    ("petekSayisi", "Petek", 9),
    ("kollektorSayisi", "Kollektör", 10),
    ("sprinkSayisi", "Sprink", 11),
    ("yanginDolabiSayisi", "Yangın dolabı", 12),
    ("ibaSayisi", "İtfaiye bağlantı ağzı", 13),
    ("araIstasyonMetraji", "Ara istasyon", 14),
    ("karotDeligiSayisi", "Karot deliği", 15),
]


def normalize_text(value):
    text = str(value or "").strip().upper()
    return (
        text.replace("İ", "I")
        .replace("Ş", "S")
        .replace("Ğ", "G")
        .replace("Ü", "U")
        .replace("Ö", "O")
        .replace("Ç", "C")
        .replace("  ", " ")
    )


def number(value):
    if value in (None, ""):
        return 0
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return 0
    return int(parsed) if parsed.is_integer() else parsed


def slug_item(category_id, index):
    return f"{category_id}_{index:02d}"


def read_templates(workbook):
    sheet = workbook["HAKEDİŞ DENEMESİ"]
    categories = []
    current = None

    for row_index in range(5, sheet.max_row + 1):
        raw_group = sheet.cell(row_index, 2).value
        item_name = sheet.cell(row_index, 3).value
        percentage = sheet.cell(row_index, 7).value

        if raw_group:
            config = CATEGORY_CONFIG.get(normalize_text(raw_group))
            if config:
                current = {**config, "items": [], "sourceRow": row_index}
                categories.append(current)

        if current and item_name and isinstance(percentage, (int, float)):
            current["items"].append(
                {
                    "id": slug_item(current["id"], len(current["items"]) + 1),
                    "name": str(item_name).strip(),
                    "sectionPercentage": round(float(percentage), 8),
                    "sourceRow": row_index,
                }
            )

    return categories


def read_legacy_coordinates(path):
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    return {str(item.get("code", "")).strip(): item.get("coordinates", []) for item in data.get("buildings", [])}


def read_buildings(workbook, coordinates_by_code):
    sheet = workbook["BİNALAR"]
    buildings = []

    for row_index in range(3, sheet.max_row + 1):
        serial = sheet.cell(row_index, 1).value
        code = sheet.cell(row_index, 2).value
        name = sheet.cell(row_index, 4).value
        if not isinstance(serial, (int, float)) or not isinstance(code, str) or not name:
            continue

        clean_code = code.strip()
        quantities = {key: number(sheet.cell(row_index, column + 1).value) for key, _, column in QUANTITY_FIELDS}
        automatic_drivers = {
            config["id"]: quantities[config["driverKey"]]
            for config in CATEGORY_CONFIG.values()
            if config["automaticWeight"]
        }
        automatic_total = sum(automatic_drivers.values())
        automatic_weights = {
            config["id"]: (
                automatic_drivers.get(config["id"], 0) / automatic_total
                if config["automaticWeight"] and automatic_total > 0
                else 0
            )
            for config in CATEGORY_CONFIG.values()
        }
        buildings.append(
            {
                "id": clean_code,
                "serial": int(serial),
                "code": clean_code,
                "name": str(name).strip(),
                "lineColor": str(sheet.cell(row_index, 3).value or "BELİRSİZ").strip(),
                "buildingCount": number(sheet.cell(row_index, 5).value) or 1,
                "quantities": quantities,
                "automaticCategoryWeights": automatic_weights,
                "coordinates": coordinates_by_code.get(clean_code, []),
                "sourceRow": row_index,
            }
        )

    return buildings


def main():
    parser = argparse.ArgumentParser(description="Excel tabanlı izole hakediş deneme verisini üretir.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument(
        "--legacy-data",
        type=Path,
        default=Path("src/deneme-dashboard/legacy/siteData.snapshot.json"),
    )
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, data_only=True, read_only=False)
    coordinates_by_code = read_legacy_coordinates(args.legacy_data)
    categories = read_templates(workbook)
    buildings = read_buildings(workbook, coordinates_by_code)

    payload = {
        "schemaVersion": 1,
        "source": {
            "file": args.workbook.name,
            "buildingSheet": "BİNALAR",
            "templateSheet": "HAKEDİŞ DENEMESİ",
        },
        "quantityFields": [{"key": key, "label": label} for key, label, _ in QUANTITY_FIELDS],
        "categories": categories,
        "buildings": buildings,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"{len(buildings)} bina, {len(categories)} kategori ve {sum(len(c['items']) for c in categories)} iş kalemi yazıldı.")


if __name__ == "__main__":
    main()
